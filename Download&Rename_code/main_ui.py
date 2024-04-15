import shutil
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import untitled

import openpyxl
import requests
from urllib.parse import urlparse
import os
import time

# 调用函数并传入 Excel 文件路径
excel_file = '1.xlsx'  # 修改为你的 Excel 文件路径

class MyWindow(untitled.Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.Rename)

    def Rename(self):

        self.label_2.setText("生成中...")
        QApplication.processEvents()
        self.Team_name = self.lineEdit.text()
        self.Times = self.lineEdit_2.text()

        self.Team_name_2 = self.lineEdit_3.text()
        self.Times_2 = self.lineEdit_4.text()

        print(self.Team_name)
        print(self.Times)
        print(self.Team_name_2)
        print(self.Times_2)

        self.Download_Rename()

    def Download_Rename(self):
        # # 如果文件夹1和文件夹2存在，则删除它们
        # if os.path.exists('文件夹1'):
        #     print("清空文件夹1")
        #     self.label_2.setText("清空文件夹1")
        #     QApplication.processEvents()
        #     for root, dirs, files in os.walk('文件夹1'):
        #         for file in files:
        #             os.remove(os.path.join(root, file))
        # else:
        #     os.makedirs('文件夹1', exist_ok=True)
        #     print("创建文件夹1")
        #     self.label_2.setText("创建文件夹1")
        #     QApplication.processEvents()
        # time.sleep(1)
        # if os.path.exists('文件夹2'):
        #     print("清空文件夹2")
        #     self.label_2.setText("清空文件夹2")
        #     QApplication.processEvents()
        #     for root, dirs, files in os.walk('文件夹2'):
        #         for file in files:
        #             os.remove(os.path.join(root, file))
        # else:
        #     os.makedirs('文件夹2', exist_ok=True)
        #     print("创建文件夹2")
        #     self.label_2.setText("创建文件夹2")
        #     QApplication.processEvents()

        # 设置一个常见的用户代理
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
        }

        # 打开 Excel 文件
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active


        
        # 分三种情况读取 Excel 文件
        # 第一种情况self.Team_name  self.Times  self.Team_name_2  self.Times_2均不为空的时候
        if (self.Team_name or self.Times) and (self.Team_name_2 or self.Times_2):

            # 如果文件夹1和文件夹2存在，则删除它们
            if os.path.exists('文件夹1'):
                print("清空文件夹1")
                self.label_2.setText("清空文件夹1")
                QApplication.processEvents()
                for root, dirs, files in os.walk('文件夹1'):
                    for file in files:
                        os.remove(os.path.join(root, file))
            else:
                os.makedirs('文件夹1', exist_ok=True)
                print("创建文件夹1")
                self.label_2.setText("创建文件夹1")
                QApplication.processEvents()
            time.sleep(1)
            if os.path.exists('文件夹2'):
                print("清空文件夹2")
                self.label_2.setText("清空文件夹2")
                QApplication.processEvents()
                for root, dirs, files in os.walk('文件夹2'):
                    for file in files:
                        os.remove(os.path.join(root, file))
            else:
                os.makedirs('文件夹2', exist_ok=True)
                print("创建文件夹2")
                self.label_2.setText("创建文件夹2")
                QApplication.processEvents()

            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                name = row[0]
                url1 = row[1]
                url2 = row[2]

                parsed_url1 = urlparse(url1)
                file_name1 = os.path.basename(parsed_url1.path)

                parsed_url2 = urlparse(url2)
                file_name2 = os.path.basename(parsed_url2.path)

                response1 = requests.get(url1, headers=headers)

                if response1.status_code == 200:
                    with open(os.path.join('文件夹1', file_name1), 'wb') as file:
                        file.write(response1.content)

                    _, file_extension1 = os.path.splitext(file_name1)
                    new_file_name1 = f"{self.Team_name}{name}{self.Times}{file_extension1}"

                    os.rename(os.path.join('文件夹1', file_name1), os.path.join('文件夹1', new_file_name1))

                    print(f"文件夹1 {file_name1} 下载并重命名为 {new_file_name1} 完成.")
                    self.label_2.setText(f"{new_file_name1} 完成")
                    QApplication.processEvents()
                else:
                    print(f"文件夹1 {file_name1} 下载失败. HTTP状态码：{response1.status_code}")

                response2 = requests.get(url2, headers=headers)

                if response2.status_code == 200:
                    with open(os.path.join('文件夹2', file_name2), 'wb') as file:
                        file.write(response2.content)

                    _, file_extension2 = os.path.splitext(file_name2)
                    new_file_name2 = f"{self.Team_name_2}{name}{self.Times_2}{file_extension2}"

                    os.rename(os.path.join('文件夹2', file_name2), os.path.join('文件夹2', new_file_name2))

                    print(f"文件夹2 {file_name2} 下载并重命名为 {new_file_name2} 完成.")
                    self.label_2.setText(f"{new_file_name2} 完成")
                    QApplication.processEvents()
                else:
                    print(f"文件夹2 {file_name2} 下载失败. HTTP状态码：{response2.status_code}")

        # 第二种情况 self.Team_name  self.Times不为空 self.Team_name_2  self.Times_2 为空的时候
        elif (self.Team_name or self.Times) and not self.Team_name_2 and not self.Times_2:

            if os.path.exists('文件夹1'):
                print("清空文件夹1")
                self.label_2.setText("清空文件夹1")
                QApplication.processEvents()
                for root, dirs, files in os.walk('文件夹1'):
                    for file in files:
                        os.remove(os.path.join(root, file))
            else:
                os.makedirs('文件夹1', exist_ok=True)
                print("创建文件夹1")
                self.label_2.setText("创建文件夹1")
                QApplication.processEvents()
            time.sleep(1)

            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                name = row[0]
                url1 = row[1]

                parsed_url1 = urlparse(url1)
                file_name1 = os.path.basename(parsed_url1.path)

                response1 = requests.get(url1, headers=headers)

                if response1.status_code == 200:
                    with open(os.path.join('文件夹1', file_name1), 'wb') as file:
                        file.write(response1.content)

                    _, file_extension1 = os.path.splitext(file_name1)
                    new_file_name1 = f"{self.Team_name}{name}{self.Times}{file_extension1}"

                    os.rename(os.path.join('文件夹1', file_name1), os.path.join('文件夹1', new_file_name1))

                    print(f"文件夹1 {file_name1} 下载并重命名为 {new_file_name1} 完成.")
                    self.label_2.setText(f"{new_file_name1} 完成")
                    QApplication.processEvents()
                else:
                    print(f"文件夹1 {file_name1} 下载失败. HTTP状态码：{response1.status_code}")
        # 第三种情况 self.Team_name  self.Times为空 self.Team_name_2  self.Times_2不为空的时候
        elif not self.Team_name and not self.Times and (self.Team_name_2 or self.Times_2):
            # 如果文件夹2存在，则删除它
            if os.path.exists('文件夹2'):
                print("清空文件夹2")
                self.label_2.setText("清空文件夹2")
                QApplication.processEvents()
                for root, dirs, files in os.walk('文件夹2'):
                    for file in files:
                        os.remove(os.path.join(root, file))
            else:
                os.makedirs('文件夹2', exist_ok=True)
                print("创建文件夹2")
                self.label_2.setText("创建文件夹2")
                QApplication.processEvents()
            time.sleep(1)

            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                name = row[0]
                url2 = row[2]  

                parsed_url2 = urlparse(url2)
                file_name2 = os.path.basename(parsed_url2.path)

                response2 = requests.get(url2, headers=headers)

                if response2.status_code == 200:
                    with open(os.path.join('文件夹2', file_name2), 'wb') as file:
                        file.write(response2.content)

                    _, file_extension2 = os.path.splitext(file_name2)
                    new_file_name2 = f"{self.Team_name_2}{name}{self.Times_2}{file_extension2}"

                    os.rename(os.path.join('文件夹2', file_name2), os.path.join('文件夹2', new_file_name2))

                    print(f"文件夹2 {file_name2} 下载并重命名为 {new_file_name2} 完成.")
                    self.label_2.setText(f"{new_file_name2} 完成")
                    QApplication.processEvents()
                else:
                    print(f"文件夹2 {file_name2} 下载失败. HTTP状态码：{response2.status_code}")
        
        




        # for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):  # 从第二行开始读取
        #     name = row[0]  # 获取名字
        #     url1 = row[1]  # 获取超链接1
        #     url2 = row[2]  # 获取超链接2

        #     # 获取文件名1
        #     parsed_url1 = urlparse(url1)
        #     file_name1 = os.path.basename(parsed_url1.path)

        #     # 获取文件名2
        #     parsed_url2 = urlparse(url2)
        #     file_name2 = os.path.basename(parsed_url2.path)

        #     # 下载文件1
        #     response1 = requests.get(url1, headers=headers)

        #     # 检查响应状态1
        #     if response1.status_code == 200:
        #         # 写入文件1
        #         with open(os.path.join('文件夹1', file_name1), 'wb') as file:
        #             file.write(response1.content)

        #         # 构建新文件名1
        #         _, file_extension1 = os.path.splitext(file_name1)
        #         new_file_name1 = f"{self.Team_name}{name}{self.Times}{file_extension1}"

        #         # 重命名文件1
        #         os.rename(os.path.join('文件夹1', file_name1), os.path.join('文件夹1', new_file_name1))

        #         print(f"文件夹1 {file_name1} 下载并重命名为 {new_file_name1} 完成.")
        #         self.label_2.setText(f"{new_file_name1} 完成")
        #         QApplication.processEvents()
        #     else:
        #         print(f"文件夹1 {file_name1} 下载失败. HTTP状态码：{response1.status_code}")

        #     # 下载文件2
        #     response2 = requests.get(url2, headers=headers)

        #     # 检查响应状态2
        #     if response2.status_code == 200:
        #         # 写入文件2
        #         with open(os.path.join('文件夹2', file_name2), 'wb') as file:
        #             file.write(response2.content)

        #         # 构建新文件名2
        #         _, file_extension2 = os.path.splitext(file_name2)
        #         new_file_name2 = f"{self.Team_name_2}{name}{self.Times_2}{file_extension2}"

        #         # 重命名文件2
        #         os.rename(os.path.join('文件夹2', file_name2), os.path.join('文件夹2', new_file_name2))

        #         print(f"文件夹2 {file_name2} 下载并重命名为 {new_file_name2} 完成.")
        #         self.label_2.setText(f"{new_file_name2} 完成")
        #         QApplication.processEvents()
        #     else:
        #         print(f"文件夹2 {file_name2} 下载失败. HTTP状态码：{response2.status_code}")

        self.label_2.setText("生成完成")
        QApplication.processEvents()


if __name__ == '__main__':
    app = QApplication(sys.argv)

    w = MyWindow()
    w.show()

    app.exec()